import openpyxl
from openpyxl.utils import get_column_letter
import config

wb = openpyxl.load_workbook(config.input_file)
sheet = wb.active

W_values = [cell.value for cell in sheet ['W']] #имя виртуальной машины
Z_values = [cell.value for cell in sheet ['Z']] #ядра цп
AD_values = [cell.value for cell in sheet ['AD']] #тип дисков
AU_values = [cell.value for cell in sheet ['AU']] #размер дисков мебибайт
V_values = [cell.value for cell in sheet ['V']] #кол-во ОЗУ мебибайт
J_values = [cell.value for cell in sheet ['J']] #версия ОС
L_values = [cell.value for cell in sheet ['L']] #внутренний IP-адрес
AC_values = [cell.value for cell in sheet ['AC']]#состояние машины
kill_element = 0
#удаляем все вм, которые находятся не в состоянии POWERED_ON
for i in range(1, len(AC_values)):
    if AC_values[i] != "POWERED_ON":
        del_element = i - kill_element
        W_values.pop(del_element)
        Z_values.pop(del_element)
        AD_values.pop(del_element)
        AU_values.pop(del_element)
        V_values.pop(del_element)
        J_values.pop(del_element)
        L_values.pop(del_element)
        kill_element = kill_element + 1

wb = openpyxl.load_workbook(config.output_file)
sheet = wb.active

#цикл для очистки листа
for row in sheet.iter_rows():
    for cell in row:
        # Очисти значение ячейки
        cell.value = None

sheet['A1'] = 'Имя VM'
sheet['B1'] = 'Колличество ядер, шт'
sheet['C1'] = 'Объем RAM, Гбайт'
sheet['D1'] = 'Объем HHD (SSD), Гбайт'
sheet['E1'] = 'Объем HHD (fast), Гбайт'
sheet['F1'] = 'Объем HHD (slow), Гбайт'
sheet['G1'] = 'IP-адрес (внутренний)'
sheet['H1'] = 'ОС'

#добавляем в таблицу имя вм и кол-во ядров цп, IP, OC потому что их не надо считать
for i in range(1, len(W_values)):
    sheet[f'A{i+1}'] = W_values[i]
    sheet[f'B{i+1}'] = Z_values[i]
    sheet[f'H{i+1}'] = J_values[i]
    sheet[f'G{i+1}'] = L_values[i]

#добавляем в таблицу оперативную память, предварительно переведя её в гб
for i in range(1, len(V_values)):
    temp = V_values[i]
    ram = temp/1024
    sheet[f'C{i+1}'] = ram

#добавляем диски, в зависимости от их назначения
for i in range(1, len(AU_values)):
    temp = AU_values[i]-V_values[i]
    disksize = temp/1024
    disksize = round(disksize)
    #добавленее HDD (fast)
    if AD_values[i] == 'FAST':
        sheet[f'E{i+1}'] = disksize
    #добавленее HDD (slow)
    elif AD_values[i] == 'STANDARD':
        sheet[f'F{i+1}'] = disksize
    #добавленее SSD
    elif AD_values[i] == 'Ultrafast_02':
        sheet[f'D{i+1}'] = disksize

    elif AD_values[i] == 'rmis-ssd':
        sheet[f'D{i+1}'] = disksize

#добавляем фильтрацию в столбцах
FullRange = "A1:" + get_column_letter(sheet.max_column) \
            + str(sheet.max_row)
sheet.auto_filter.ref = FullRange

wb.save(config.output_file)
print("Файл подготовлен")